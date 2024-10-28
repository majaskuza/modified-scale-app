import pandas as pd
from datetime import datetime
import requests
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO


def create_general_info_spreadsheet(url, cookies):
    # Make the request with cookies to get the Excel file
    response = requests.get(url, cookies=cookies)

    # Check if the request was successful
    if response.status_code != 200:
        raise Exception(f"Failed to retrieve file: {response.status_code}")

    # Check the content type
    content_type = response.headers.get('Content-Type')
    print(f"Content-Type: {content_type}")

    if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" not in content_type:
        print("The response is not an Excel file.")
        print(f"Response Text: {response.text[:100]}")  # Print first 100 characters of the response text
        return

    # Load the workbook from the response content
    workbook = load_workbook(filename=BytesIO(response.content))

    # Create a new worksheet or get the existing one
    new_sheet_name = "General Info"
    if new_sheet_name in workbook.sheetnames:
        worksheet = workbook[new_sheet_name]
    else:
        worksheet = workbook.create_sheet(title=new_sheet_name)

    # Create the data and add it to the new worksheet
    general_info_data = {
        'Cage': [],
        'ID': [],
        'Name': [],
        'Start weight': [],
        '80% weight': []
    }
    general_info_df = pd.DataFrame(general_info_data)

    # Write the DataFrame to the new worksheet
    for r in dataframe_to_rows(general_info_df, index=False, header=True):
        worksheet.append(r)

    # Save the modified workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Prepare to upload the modified workbook back to SharePoint
    headers = {
        'Accept': 'application/json;odata=verbose',
        'X-RequestDigest': requests.post(f"{url}/_api/contextinfo", cookies=cookies).json()['d']['GetContextWebInformation']['FormDigestValue'],
        'IF-MATCH': '*',
        'X-HTTP-Method': 'MERGE'
    }

    # Use the same URL for uploading; make sure it's the correct endpoint for the file
    response = requests.post(url, headers=headers, data=output.read(), cookies=cookies)

    # Check if the upload was successful
    if response.status_code == 200:
        print("Successfully added new worksheet to the file!")
    else:
        print(f"Failed to upload modified file: {response.status_code}, {response.text}")



#function to add the weight information from a single session into the excel spreedsheet
def add_recording_entry(url, subjid, weight, animal_fed_at='', training_end='', food_amount=''):
    timestamp = datetime.now().strftime("%H:%M")  
    percent = 0.8 * weight  
    new_entry = {
        'Date': datetime.now().strftime("%Y-%m-%d"),  
        'Subjid': subjid,
        'Training Start': timestamp,
        'Training End': training_end,
        'Animal Fed At': animal_fed_at,
        'Food Amount': food_amount,
        'Percent': percent,
        'Weight': weight
    }
    try:
        df = pd.read_excel(url)
    except FileNotFoundError:
        df = pd.DataFrame(columns=new_entry.keys())  # Create an empty DataFrame with the columns
    df = df.append(new_entry, ignore_index=True)
    df.to_excel(url, index=False)


#url = "https://liveuclac.sharepoint.com/:x:/s/swcpil/Ef2jWbBmIkRKmG7a-NQwVyUBNjUFgPr-tQSDpJ5e7XOpcQ?email=maja.skuza.23%40ucl.ac.uk&e=088cEn"
#cookies = {
   # 'FedAuth': '77u/PD94bWwgdmVyc2lvbj0iMS4wIiBlbmNvZGluZz0idXRmLTgiPz48U1A+VjEzLDBoLmZ8bWVtYmVyc2hpcHwxMDAzMjAwMmRiZmIyZGQ3QGxpdmUuY29tLDAjLmZ8bWVtYmVyc2hpcHx1Y3R2YXNrQHVjbC5hYy51aywxMzM3MzcxODI2MTAwMDAwMDAsMTMzNzM1NDQyOTMwMDAwMDAwLDEzMzc1MDIwNzQ5NTcwOTIyMCwxMjguNDAuMTc3LjEwLDY3LDQ2ZDRlYzMzLWMwNTMtNGZhOC05Y2UxLTBjYTdjOTA0NzY1YiwsMDJmZDY5MTUtZTJjNy00NWJkLWFlYWItZDA1NTc5ZjA4YzE1LDZjNjM1ZGExLTEwMzYtYTAwMC00M2I5LWJmM2UyM2M3NGE3YywyZjUxNWVhMS0zMDZmLWEwMDAtMmZkOC1iNzc2NDdhZjdiZDksLDAsMTMzNzQ1OTEzNzcwNDE3NDkzLDEzMzc0ODQ2OTc3MDQxNzQ5MywsLGV5SjRiWE5mWTJNaU9pSmJYQ0pEVURGY0lsMGlMQ0o0YlhOZmMzTnRJam9pTVNJc0luQnlaV1psY25KbFpGOTFjMlZ5Ym1GdFpTSTZJblZqZEhaaGMydEFkV05zTG1GakxuVnJJaXdpZFhScElqb2lUVmt6YUUwMVNqZE5hMmxqVjFONVkzTjRPVXBCUVNKOSwyNjUwNDY3NzQzOTk5OTk5OTk5LDEzMzc0MzM5NDM3MDAwMDAwMCwzNDM0ZjU5My0xYjM4LTQ4MzQtYjRhZi0yYjhkNzk2Njk2OWMsLCwsLCwxMTUyOTIxNTA0NjA2ODQ2OTc2LCwxODk5NDMsdVhlaFFKUGxlVmpOQ2Jha1VoR0Q2SXlGUVFrLERMRTZuamdTd0RDVlRqWUxNZW1Wb3FqUVF1VnNEODNYbTRMTUw3anM5WnNrazEyWlM2RVpHK0d2L2NYa1FpdWZ3SFRMR1BxZXZJOTlWY3V3NUpubkhkY2VpUWMzS1QwdCtRbFRHNGkzTDZ2WUI3bXJEczZwL2hXVGpYU1VCZm5MbFl0U2VCSDlaZm5FWmNVQmpSZjFxeU5ncEFPS2tlcUd5NStIWnJaZktRQzlNSUtvN3pDcElIMlA2alE4VEJZTWQvVEpLRGlwbFAzc0o4aHB6VWRYNUJlTi9ZMDVXU0ZySHpSUTFzOGxkbUkrQnFSNWNSb21pRndrUFl1OU9hT1hHMGFMQUUwNVE1Z0hlb2Z3YmhZY0JFd0pKWkJHdEF3ZGVOWEdudWkweWlWUStNNGw4MWRwSWU2am1kVFJWZG1xZ2VVYWN6MG1BMzIyMXN1VWE0K1ViUT09PC9TUD4=',  # Replace with your actual FedAuth value
   # 'rtFa': 'O0LCC18/78wrfbORuFTWLXEwq6x1KsrK9mtnfbU2uqMmNDZkNGVjMzMtYzA1My00ZmE4LTljZTEtMGNhN2M5MDQ3NjViIzEzMzc0MzM5NDM3NDcxODE5MiM2YzYzNWRhMS0xMDM2LWEwMDAtNDNiOS1iZjNlMjNjNzRhN2MjdWN0dmFzayU0MHVjbC5hYy51ayMxODk5NDMjYjNJUWg0NWt5YWZ4dmZKVE5TUDNSc3VvT0ZNcX1pKfpEqY9FW2YVjoPkkaKOj/VtzBqZSfb7tZQMUiqC67tgt4iudWt/UUHMUa1KjdoyAKQzObaZfwcFLhnzZib1gTY2Ru5KD061ibeCaEAOduM6H8B3JqsmMsoE8G09JIKmlTZJF4gCD1wT+v0A6qMwhqhe3gCvl0580ZQt8t9Qry69u6DmUkYceq+7IAr13TzU1LYVQ9VquOzIkdwk0RU/7DDHjLw/WsjJ1u6elCpch0MFOYkrLTnYveL+mNcCv73w86CaTPSC6N9MckPz26fAAqEad5ohVqKTNZtrBoH7nkdgHut5xNxwxGdUV2TgZ7HYxAltSagUc/+AGIKuXbQAAAA=' }

#example usage of above functions
#create_general_info_spreadsheet(url, cookies)
